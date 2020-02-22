#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import openpyxl
import matplotlib.pyplot as plt
import numpy as np

#用于matplotlib显示中文
plt.rcParams['font.sans-serif'] = ['SimHei']

fpath =  './股票.xlsx'
stock_SIndex_col = 2 #B
stock_SIndex_row = 2
stock_EIndex_row = 13 #股票数
stock_wave_SIndex_col = 5 #E 波动起始位子
stock_wave_EIndex_col = 34 #波动结束位子
RECODE_CYCLE = 30
color = ['yellow','red','coral','aliceblue','antiquewhite','black'
         ,'blue','cadetblue','chocolate','cyan','darkgray','darkorchid'
         ,'darkslategray','floralwhite','gold']

#with openpyxl.load_workbook(fpath) as wb:
wb = openpyxl.load_workbook(fpath)
sheets = wb.sheetnames
#print(sheets, type(sheets))

ws = wb[wb.sheetnames[1]]
print(type(ws.cell(stock_SIndex_col,stock_SIndex_row).value))
LL = [[] for i in range(stock_EIndex_row - stock_SIndex_row + 1)]
#一行打印一次
for column,num in zip(ws.iter_rows(min_row=stock_SIndex_row,max_row=stock_EIndex_row,
                           min_col=stock_wave_SIndex_col, max_col = stock_wave_EIndex_col, values_only=True),
                       range(stock_EIndex_row - stock_SIndex_row + 1)):
    plt.figure(num=3,figsize=(10,5))
    for i in range(RECODE_CYCLE):
        if not column[i] == None:
            #print(type(column[i]))
            if type(column[i]) == int or type(column[i]) == float:
                #print(column[i])
                LL[num].append(column[i])
            elif(type(column[i]) == str):
                LL[num].append(column[i].split('/')[0])
                #print(LL)
            else:print(type(column[i]),column[i])
        else:
            LL[num].append(0)
        pass
    i = np.arange(0,RECODE_CYCLE)
    LL[num] = list(map(float,LL[num]))
    print(type(LL[num]),LL[num])
    plt.xlabel("时间(天)")
    plt.ylabel('波幅')
    plt.plot(i, LL[num], color=color[num], linewidth=1.0, linestyle='-',label = ws.cell(num+stock_SIndex_col,stock_SIndex_row).value)
    pass
plt.legend()
plt.show()

#wb.save(fpath)  #read_only state not need